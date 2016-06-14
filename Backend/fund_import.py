# -*- coding: utf-8 -*-

import logging
import re
import datetime
import mysql.connector
from mysql.connector import errorcode

import sys
reload(sys)
sys.setdefaultencoding('utf8')

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

import random
from retrying import retry

from openpyxl import load_workbook


def import_zyq_score(filename):

    result = []

    wb = load_workbook(filename)
    ws = wb['Mthly ZYQ Score']

    # for row_index, row in enumerate(ws.rows):
    #     for col_index, cell in enumerate(row):
    #         print row_index, col_index, cell.value

    rownum = 0
    for row in ws.rows:
        rownum += 1
        # print "processing row", rownum

        if rownum == 1:
            pass
        elif rownum == 2:
            row_day = row
        else:
            for i in range(4, 182):
                perf_year = row_day[i].value.split("/")[2].strip()
                perf_month = row_day[i].value.split("/")[0].strip()
                result.append([row[1].value,
                               row[2].value,
                               row[3].value,
                               perf_year,
                               perf_month,
                               int(perf_year)*12+int(perf_month),
                               row[i].value
                               ])

    # delete all duplicated records:
    cursor = cnx.cursor()
    cursor.execute("""DELETE FROM t_fund_performance where perf_type='zyq_score'""")
    logging.info(unicode(cursor.rowcount) + ' zyq_score data deleted')

    add_product = ("""INSERT INTO t_fund_performance(fund_name, isin_code, currency, perf_year, perf_month, months,
                      perf_type, performance)
                      VALUES (%s, %s, %s, %s, %s, %s, 'zyq_score', %s)""")

    # sqlnum = 1
    for pd in result:
    #     print "insert row", sqlnum
        cursor.execute(add_product, pd)
    #     sqlnum+=1

    logging.info(unicode(len(result)) + ' zyq_score data imported')
    cursor.close()


def import_fund_performance(filename):

    result = []

    wb = load_workbook(filename)
    ws = wb['Mthly Performance']

    # for row_index, row in enumerate(ws.rows):
    #     for col_index, cell in enumerate(row):
    #         print row_index, col_index, cell.value

    rownum = 0
    for row in ws.rows:
        rownum += 1
        # print "processing row", rownum

        if rownum == 1:
            pass
        elif rownum == 2:
            row_day = row
        else:
            for i in range(4, 248):
                perf_year = row_day[i].value.split("/")[2].strip()
                perf_month = row_day[i].value.split("/")[0].strip()
                result.append([row[1].value,
                               row[2].value,
                               row[3].value,
                               perf_year,
                               perf_month,
                               int(perf_year)*12+int(perf_month),
                               row[i].value
                               ])

    # delete all duplicated records:
    cursor = cnx.cursor()
    cursor.execute("""DELETE FROM t_fund_performance where perf_type='mthly_performance'""")
    logging.info(unicode(cursor.rowcount) + ' mthly_performance data deleted')

    add_product = ("""INSERT INTO t_fund_performance(fund_name, isin_code, currency, perf_year, perf_month, months,
                       perf_type, performance)
                      VALUES (%s, %s, %s, %s, %s, %s, 'mthly_performance', %s)""")

    # sqlnum = 1
    for pd in result:
    #     print "insert row", sqlnum
        cursor.execute(add_product, pd)
    #     sqlnum+=1

    logging.info(unicode(len(result)) + ' mthly_performance data imported')
    cursor.close()


def decode(*arguments):
    """Compares first item to subsequent item one by one.

    If first item is equal to a key, returns the corresponding value (next item).
    If no match is found, returns None, or, if default is omitted, returns None.

    example usage:
    return_value = decode('b', 'a', 1, 'b', 2, 3)
    var = 'list'
    return_type = decode(var, 'tuple', (), 'dict', {}, 'list', [], 'string', '')
    """
    if len(arguments) < 3:
        raise TypeError, 'decode() takes at least 3 arguments (%d given)' % (len(arguments))
    de_dict = list(arguments[1:])
    if arguments[0] in de_dict:
        index = de_dict.index(arguments[0]);
        if index % 2 == 0 and len(de_dict) > index+1:
            return de_dict[index+1]
        return de_dict[-1]
    elif len(de_dict) % 2 != 0:
        return de_dict[-1]


if __name__ == '__main__':
    DB_NAME = 'zyq'

    try:
        # setup console parameters

        cnx = mysql.connector.connect(host='localhost', user='zyq', password='zyq', database=DB_NAME)
        # cnx = mysql.connector.connect(host='localhost', user='zyq', password='zyq', database=DB_NAME)
        logging.info('MYSQL connected.')

        import_zyq_score(sys.argv[1])
        import_fund_performance(sys.argv[1])

        cnx.commit()
        cnx.close()

    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)
            raise
