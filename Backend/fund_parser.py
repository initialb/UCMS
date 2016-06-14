# -*- coding: utf-8 -*-

import math
import codecs
import argparse
import logging
import traceback
import re
import requests
import bs4
import json
import StringIO
import time
import openpyxl
import mysql.connector
from mysql.connector import errorcode
from multiprocessing import Pool
from decimal import Decimal
from butils import decode
from butils import fix_json
from butils import ppprint
from datetime import datetime
from HTMLParser import HTMLParser
from butils.pprint import pprint

import sys, getopt
reload(sys)
sys.setdefaultencoding('utf8')

import random
from retrying import retry

TIMEOUT = 60
LOCALTIME = time.strftime('%Y%m%d', time.localtime(time.time()))
LOGNAME = 'log/fund_parser_' + LOCALTIME + '.log'

# initialize root logger to write verbose log file (inculding logs from all called packages like requests etc.)
logging.basicConfig(level=logging.DEBUG,
                    filename="log/fund_parser_" + LOCALTIME + ".verbose.log",
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# initialize a local logger
logger_local = logging.getLogger("ucms.birdie.parser.fund")
logger_local.setLevel(logging.DEBUG)

# initialize a local logger file handler
logger_local_fh = logging.FileHandler(LOGNAME)
logger_local_fh.setLevel(logging.INFO)
logger_local_fh.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

# initialize a local logger console handler
logger_local_ch = logging.StreamHandler()
logger_local_ch.setLevel(logging.INFO)
logger_local_ch.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

# add handlers to local logger
logger_local.addHandler(logger_local_fh)
logger_local.addHandler(logger_local_ch)

VERSION = 0.5

# FundSuperMart
# 请求格式：http post
# 返回格式：html


def get_FSM_fund_product():
    try:
        index_url = 'http://www.fundsupermart.com.hk/hk/main/fundinfo/generateTable.svdo?lang=zh'
        logging.info("Retrieving " + index_url + " ...")

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency"})

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        fund_data = []

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FUNDLIST"
        ws_title = ["prod_code",
                    "prod_name",
                    "risk_rating",
                    "currency",
                    "latest_nav_price",
                    "3year_risk_return_ratio",
                    "sharpe_ratio",
                    "valuation_date",
                    "cp_1m",
                    "cp_3m",
                    "cp_6m",
                    "cp_ytd",
                    "cp_1y",
                    "cp_3y",
                    "cp_5y",
                    "cp_since_launch",
                    "launch_date",
                    "cyp_1",
                    "cyp_2",
                    "cyp_3",
                    "cyp_4",
                    "cyp_5"]
        ws.append(ws_title)
        for fund in fund_data:
            ws.append(fund)
        wb.save(output_file)

        # delete all duplicated records:
        cursor = cnx.cursor()
        cursor.execute("""DELETE FROM t_fund_product WHERE data_source='FSM' and date(update_time)=curdate()""")
        logging.info(unicode(cursor.rowcount) + ' FSM funds deleted')

        add_product = ("""INSERT INTO t_fund_product(prod_code, prod_name, risk_rating, currency, latest_nav_price,
                          3year_risk_return_ratio, sharpe_ratio, valuation_date, cp_1m, cp_3m, cp_6m, cp_ytd,
                          cp_1y, cp_3y, cp_5y, cp_since_launch, launch_date, cyp_1, cyp_2, cyp_3, cyp_4, cyp_5,
                          data_source, update_time)
                          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'FSM', now())""")

        for fund in fund_data:
            cursor.execute(add_product, fund)

        logging.info(unicode(len(fund_data)) + ' FSM funds imported')
        cursor.close()

    except:
        raise


def get_FSM_fund_classified_product():
    try:
        index_url = 'http://www.fundsupermart.com.hk/hk/main/fundinfo/generateTable.svdo'
        logging.info("Retrieving " + index_url + " ...")
        fund_data = []
        fund_data_or = []

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "AI"})
            # 另类基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "另类基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "MB"})
            # 均衡基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "均衡基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "EG"})
            # 股票基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "股票基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)


        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "FI"})
            # 定息基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "定息基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "MA"})
            # 混合资产基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "混合资产基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)


        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency", "sectormaincode": "DG"})
            # 货币市场基金

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "货币市场基金",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data)


        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.post(index_url, timeout=20, data={"baseCur": "fundCurrency"})

        response = request_content()

        buf = StringIO.StringIO(response.text)
        f = StringIO.StringIO()

        is_first_tr = True
        line = True

        # 处理返回结果中<tr>未闭合的问题, 否则soup分析出错.
        while line:
            line = buf.readline()
            if " --------------Start displaying fund list-------------------  " in line:
                while True:
                    line = buf.readline()
                    if "<tr" in line:
                        # 第一行前不增加</br>
                        if is_first_tr is True:
                            is_first_tr = False
                        else:
                            f.write(" "*17*4 + "</tr>\n")
                    # 读取到table最后一个tr
                    if "</table>" in line:
                        break
                    # 跳过空行
                    if line.strip() == "":
                        continue
                    f.write(line)
                break

        buf.close()

        # 回到文件头, 读取至soup分析
        f.seek(0)
        fund_list_html = f.read()
        f.close()

        soup = bs4.BeautifulSoup(fund_list_html, "html.parser")
        fund_list = soup.find_all("tr")

        for idx, fund in enumerate(fund_list):
            tds = fund.find_all("td")
            if len(tds) == 22:
                fund_data_or.append([tds[0].input["value"],
                                  tds[1].a.font.text.strip(),
                                  "",
                                  "",
                                  "",
                                  "",
                                  tds[2].font.text.strip(),
                                  tds[3].font.text.strip(),
                                  tds[4].font.text.strip(),
                                  tds[5].font.text.strip(),
                                  tds[6].font.text.strip(),
                                  tds[7].font.text.strip(),
                                  tds[8].font.text.strip(),
                                  tds[9].font.text.strip(),
                                  tds[10].font.text.strip(),
                                  tds[11].font.text.strip(),
                                  tds[12].font.text.strip(),
                                  tds[13].font.text.strip(),
                                  tds[14].font.text.strip(),
                                  tds[15].font.text.strip(),
                                  tds[16].font.text.strip(),
                                  tds[17].font.text.strip(),
                                  tds[18].font.text.strip(),
                                  tds[19].font.text.strip(),
                                  tds[20].font.text.strip(),
                                  tds[21].font.text.strip()])
        logging.debug(fund_data_or)


        for data in fund_data_or:
            exists = 0
            for d in fund_data:
                if data[0] == d[0]:
                    exists += 1
            if exists == 0:
                fund_data.append(data)


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FUNDLIST"
        ws_title = ["prod_code",
                    "prod_name",
                    "prod_type",
                    "placeholder_1"
                    "placeholder_2"
                    "placeholder_3"
                    "risk_rating",
                    "currency",
                    "latest_nav_price",
                    "3year_risk_return_ratio",
                    "sharpe_ratio",
                    "valuation_date",
                    "cp_1m",
                    "cp_3m",
                    "cp_6m",
                    "cp_ytd",
                    "cp_1y",
                    "cp_3y",
                    "cp_5y",
                    "cp_since_launch",
                    "launch_date",
                    "cyp_1",
                    "cyp_2",
                    "cyp_3",
                    "cyp_4",
                    "cyp_5"]
        ws.append(ws_title)
        for fund in fund_data:
            ws.append(fund)
        wb.save(output_file)

        # delete all duplicated records:
        # cursor = cnx.cursor()
        # cursor.execute("""DELETE FROM t_fund_product WHERE data_source='FSM' and date(update_time)=curdate()""")
        # logging.info(unicode(cursor.rowcount) + ' FSM funds deleted')
        #
        # add_product = ("""INSERT INTO t_fund_product(prod_code, prod_name, risk_rating, currency, latest_nav_price,
        #                   3year_risk_return_ratio, sharpe_ratio, valuation_date, cp_1m, cp_3m, cp_6m, cp_ytd,
        #                   cp_1y, cp_3y, cp_5y, cp_since_launch, launch_date, cyp_1, cyp_2, cyp_3, cyp_4, cyp_5,
        #                   data_source, update_time)
        #                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        #                           %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'FSM', now())""")
        #
        # for fund in fund_data:
        #     cursor.execute(add_product, fund)

        logging.info(unicode(len(fund_data)) + ' FSM funds imported')
        cursor.close()

    except:
        raise


def get_MS_fund_page_num():
    try:
        index_url = 'http://www.hk.morningstar.com/ap/fundselect/results.aspx'

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            return requests.get(index_url, timeout=TIMEOUT)

        response = request_content()
        soup = bs4.BeautifulSoup(response.text, "html.parser")
        total_record = soup.find("span", id="MainContent_TotalResultLabel").text
        logger_local.info('[FI]Total record: %s' % total_record)
        logger_local.info('[FI]Total page: %s' % (int(total_record)/30+1))

        return int(total_record)/30+1

    except:
        raise


def get_MS_fund_product(total_page):
    index_url = 'http://www.hk.morningstar.com/ap/fundselect/results.aspx'

    cursor.execute("""
                   DELETE FROM t_fund_product WHERE data_source='MS'
                   """)
    logger_local.info('MS - ' + unicode(cursor.rowcount) + ' rows deleted')

    for page in range(total_page):
        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            return requests.post(index_url, timeout=60, data={"__EVENTTARGET": "ctl00$MainContent$AspNetPager1",
                                                              "__EVENTARGUMENT": page+1,
                                                              "__VIEWSTATE": ""})

        response = request_content()

        soup = bs4.BeautifulSoup(response.text, "html.parser")
        prod_list_tr = soup.find("table", id="MainContent_gridResult").find_all("tr")

        print "Fetching page: ", page+1
        prod_list = []

        for prod in prod_list_tr:
            if prod.has_attr("class"):
                prod_list_td = prod.find_all("td")
                prod_list.append([re.sub(r"\.\.", "", prod_list_td[0].a["href"]),
                                  prod_list_td[0].text,
                                  prod_list_td[1].text,
                                  re.sub(r"[^\d]",
                                         "",
                                         prod_list_td[2].img["src"] if prod_list_td[2].find("img") else "--"),
                                  prod_list_td[3].text,
                                  prod_list_td[4].text])

        add_product = ("""INSERT INTO t_fund_product(remark, prod_name, GIFS_type_cn, risk_rating, currency,
                          latest_nav_price, data_source, update_time)
                          VALUES (%s, %s, %s, %s, %s, %s, 'MS', now())""")

        for fund in prod_list:
            cursor.execute(add_product, fund)


def get_MS_fund_product_detail():
    index_url = 'http://www.hk.morningstar.com/ap'
    prod_list = []

    query = "select uid, prod_name, remark from t_fund_product where data_source='MS'"
    cursor.execute(query)
    for (uid, prod_name, remark) in cursor:
        prod_list.append([uid, prod_name, index_url+remark])

    for p in range(len(prod_list)):
    # for p in range(2):
        logger_local.info("(%s/%s) requesting %s [%s]" % (p+1, len(prod_list), prod_list[p][1], prod_list[p][2]))

        """
        检索概况页面信息
        """
        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            return requests.get(prod_list[p][2], timeout=TIMEOUT)

        response = request_content()
        soup = bs4.BeautifulSoup(response.text, "html.parser")

        # print response.text

        try:
            ISIN_code = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_ISINText").text
        except AttributeError, e:
            ISIN_code = None
        try:
            GIFS_type = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_GIFSText").text
        except AttributeError, e:
            GIFS_type = None
        try:
            IFA_category = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_IFASectorText").text
        except AttributeError, e:
            IFA_category = None
        try:
            industry_1 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_0").find_parent("li").span.text
        except AttributeError, e:
            industry_1 = None
        try:
            industry_pct_1 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_0").find_parent("ul").find("li", class_="regionbreakdown_col2").span.text
        except AttributeError, e:
            industry_pct_1 = None
        try:
            industry_2 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_1").find_parent("li").span.text
        except AttributeError, e:
            industry_2 = None
        try:
            industry_pct_2 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_1").find_parent("ul").find("li", class_="regionbreakdown_col2").span.text
        except AttributeError, e:
            industry_pct_2 = None
        try:
            industry_3 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_2").find_parent("li").span.text
        except AttributeError, e:
            industry_3 = None
        try:
            industry_pct_3 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_2").find_parent("ul").find("li", class_="regionbreakdown_col2").span.text
        except AttributeError, e:
            industry_pct_3 = None
        try:
            industry_4 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_3").find_parent("li").span.text
        except AttributeError, e:
            industry_4 = None
        try:
            industry_pct_4 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_3").find_parent("ul").find("li", class_="regionbreakdown_col2").span.text
        except AttributeError, e:
            industry_pct_4 = None
        try:
            industry_5 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_4").find_parent("li").span.text
        except AttributeError, e:
            industry_5 = None
        try:
            industry_pct_5 = soup.find("img", id="MainContent_QuickTakeMainContent_QuickTakeForm_StockBreakdownRepeater_SectorImage_4").find_parent("ul").find("li", class_="regionbreakdown_col2").span.text
        except AttributeError, e:
            industry_pct_5 = None

        """
        检索回报页面信息
        """
        prod_code = re.sub(r".+(?==)=", "", prod_list[p][2])
        total_return_url = "http://www.hk.morningstar.com/ap/quicktake/returns.aspx?PerformanceId=" + prod_code + "&activetab=TotalReturn"

        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            return requests.get(total_return_url, timeout=TIMEOUT)

        response = request_content()
        soup = bs4.BeautifulSoup(response.text, "html.parser")

        try:
            cp_1d = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label1").text
        except AttributeError, e:
            cp_1d = None
        try:
            cp_1w = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label6").text
        except AttributeError, e:
            cp_1w = None
        try:
            cp_1m = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label11").text
        except AttributeError, e:
            cp_1m = None
        try:
            cp_3m = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label16").text
        except AttributeError, e:
            cp_3m = None
        try:
            cp_6m = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label26").text
        except AttributeError, e:
            cp_6m = None
        try:
            cp_ytd = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label21").text
        except AttributeError, e:
            cp_ytd = None
        try:
            cp_1y = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label26").text
        except AttributeError, e:
            cp_1y = None
        try:
            cp_3y = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label31").text
        except AttributeError, e:
            cp_3y = None
        try:
            cp_5y = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label36").text
        except AttributeError, e:
            cp_5y = None
        try:
            cp_10y = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_Label41").text
        except AttributeError, e:
            cp_10y = None
        try:
            trailing_return_update_time = soup.find("span", id="MainContent_QuickTakeMainContent_QuickTakeForm_lbEffectDatem").text
            if len(trailing_return_update_time) != 10:
                trailing_return_update_time = '0001-01-01'
        except AttributeError, e:
            trailing_return_update_time = '0001-01-01'

        print ISIN_code, prod_list[p][0]

        update_product = ("""UPDATE t_fund_product SET GIFS_type=%s, IFA_category=%s, industry_1=%s, industry_pct_1=%s,
                             industry_2=%s, industry_pct_2=%s, industry_3=%s, industry_pct_3=%s, industry_4=%s,
                             industry_pct_4=%s, industry_5=%s, industry_pct_5=%s, cp_1d=%s, cp_1w=%s, cp_1m=%s,
                             cp_3m=%s, cp_6m=%s, cp_ytd=%s, cp_1y=%s, cp_3y=%s, cp_5y=%s, cp_10y=%s, prod_code=%s,
                             ISIN_code=%s, update_time=%s
                             WHERE uid=%s""")

        cursor.execute(update_product, (GIFS_type, IFA_category, industry_1, industry_pct_1, industry_2, industry_pct_2,
                                        industry_3, industry_pct_3, industry_4, industry_pct_4, industry_5,
                                        industry_pct_5, cp_1d, cp_1w, cp_1m, cp_3m, cp_6m, cp_ytd, cp_1y, cp_3y, cp_5y,
                                        cp_10y, prod_code, ISIN_code, trailing_return_update_time, prod_list[p][0]))
        cnx.commit()


def set_MS_region():
    prod_list = []
    query = "select uid, ifa_category from t_fund_product where data_source='MS'"
    cursor.execute(query)
    for (uid, ifa_category) in cursor:
        if "Equity" in ifa_category:
            region = ifa_category.replace("Equity", "").strip()
            category = "Equity"
        elif "Asset Allocation" in ifa_category:
            region = ifa_category.replace("Asset Allocation", "").strip()
            category = "Asset Allocation"
        elif "Fixed Income" in ifa_category:
            region = ifa_category.replace("Fixed Income", "").strip()
            category = "Fixed Income"
        elif "Money Market" in ifa_category:
            region = ifa_category.replace("Money Market", "").strip()
            category = "Money Market"
        elif "Smaller Companies" in ifa_category:
            region = ifa_category.replace("Smaller Companies", "").strip()
            category = "Smaller Companies"
        else:
            region = None
            category = ifa_category

        prod_list.append([uid, category, region])

    update_product = """update t_fund_product set category=%s, region=%s where uid=%s"""
    for p in prod_list:
        cursor.execute(update_product, (p[1], p[2], p[0]))


def MS_export():

    fund_list = []
    query = u"""
        SELECT
            prod_code,
            prod_name,
            region,
            category,
            ISIN_code,
            GIFS_type,
            GIFS_type_cn,
            IFA_category,
            risk_rating,
            currency,
            latest_nav_price,
            3year_risk_return_ratio,
            sharpe_ratio,
            valuation_date,
            cp_1d,
            cp_1w,
            cp_1m,
            cp_3m,
            cp_6m,
            cp_ytd,
            cp_1y,
            cp_3y,
            cp_5y,
            cp_10y,
            cp_since_launch,
            launch_date,
            cyp_1,
            cyp_2,
            cyp_3,
            cyp_4,
            cyp_5,
            industry_1,
            industry_pct_1,
            industry_2,
            industry_pct_2,
            industry_3,
            industry_pct_3,
            industry_4,
            industry_pct_4,
            industry_5,
            industry_pct_5,
            remark,
            data_source,
            update_time
        FROM
            t_fund_product"""
    cursor.execute(query)
    for (prod_code, prod_name, region, category, ISIN_code, GIFS_type, GIFS_type_cn, IFA_category, risk_rating,
         currency, latest_nav_price, year_risk_return_ratio, sharpe_ratio, valuation_date, cp_1d, cp_1w, cp_1m, cp_3m,
         cp_6m,
         cp_ytd, cp_1y, cp_3y, cp_5y, cp_10y, cp_since_launch, launch_date, cyp_1, cyp_2, cyp_3, cyp_4, cyp_5,
         industry_1, industry_pct_1, industry_2, industry_pct_2, industry_3, industry_pct_3, industry_4,
         industry_pct_4, industry_5, industry_pct_5, remark, data_source, update_time) in cursor:
        fund_list.append([prod_code, prod_name, region, category, ISIN_code, GIFS_type, GIFS_type_cn, IFA_category,
                          risk_rating, currency, latest_nav_price, year_risk_return_ratio, sharpe_ratio,
                          valuation_date, cp_1d, cp_1w, cp_1m, cp_3m, cp_6m, cp_ytd, cp_1y, cp_3y, cp_5y, cp_10y,
                          cp_since_launch,
                          launch_date, cyp_1, cyp_2, cyp_3, cyp_4, cyp_5, industry_1, industry_pct_1, industry_2,
                          industry_pct_2, industry_3, industry_pct_3, industry_4, industry_pct_4, industry_5,
                          industry_pct_5, remark, data_source, update_time])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FUNDLIST"
    ws_title = [
        "prod_code",
        "prod_name",
        "region",
        "category",
        "ISIN_code",
        "GIFS_type",
        "GIFS_type_cn",
        "IFA_category",
        "risk_rating",
        "currency",
        "latest_nav_price",
        "3year_risk_return_ratio",
        "sharpe_ratio",
        "valuation_date",
        "cp_1m",
        "cp_3m",
        "cp_6m",
        "cp_ytd",
        "cp_1y",
        "cp_3y",
        "cp_5y",
        "cp_10y",
        "cp_since_launch",
        "launch_date",
        "cyp_1",
        "cyp_2",
        "cyp_3",
        "cyp_4",
        "cyp_5",
        "industry_1",
        "industry_pct_1",
        "industry_2",
        "industry_pct_2",
        "industry_3",
        "industry_pct_3",
        "industry_4",
        "industry_pct_4",
        "industry_5",
        "industry_pct_5",
        "remark",
        "data_source",
        "update_time"]
    ws.append(ws_title)
    for fund in fund_list:
        ws.append(fund)
    wb.save(output_file)


def get_jpm_fund_product():
    root_url = 'https://www.jpmorganam.com.hk'
    index_url = 'https://www.jpmorganam.com.hk/jpm/am/zh/funds/performance'
    fund_data = []
    logging.info("Retrieving " + index_url + " ...")


    try:
        @retry(stop_max_attempt_number=10, wait_fixed=2000)
        def request_content():
            logging.info("Retrieving " + index_url + " ...")
            return requests.get(index_url, timeout=20)

        response = request_content()

        soup = bs4.BeautifulSoup(response.text, "html.parser")
        fund_list_tb = soup.find("table", id="priceColumn").find_all("tr")

        for idx, fund in enumerate(fund_list_tb):
            if idx == 0:
                pass
            else:
                fund_td = fund.find_all("td")
                if fund_td[1].a:
                    prod_name = fund_td[1].a.text.strip()
                    prod_url = root_url + fund_td[1].a["href"]
                else:
                    prod_name = fund_td[1].text.strip()
                    prod_url = ""
                fund_data.append([fund_td[0].text.strip(),  # prod code
                                  prod_name,
                                  prod_url,
                                  fund_td[2].text.strip(),  # start date
                                  fund_td[3].text.strip(),  # currency
                                  fund_td[4].text.strip(),  # cp ytd
                                  fund_td[5].text.strip(),  # cp 1y
                                  fund_td[6].text.strip(),  # cp 3y
                                  fund_td[7].text.strip(),  # cp 5y
                                  fund_td[8].text.strip()])  # cp since launch
                # logging.info(fund_data[-1])
                ppprint(fund_data[-1])

                @retry(stop_max_attempt_number=10, wait_fixed=2000)
                def request_content():
                    logging.info("Retrieving " + index_url + " ...")
                    return requests.get(prod_url,
                                        cookies=response.cookies,
                                        headers={"referer": index_url, "Host": "www.jpmorganam.com.hk"},
                                        timeout=20)

                pd_response = request_content()

                pd_soup = bs4.BeautifulSoup(pd_response.text, "html.parser")

                print pd_soup
                exit(0)


        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "FUNDLIST"
        # ws_title = ["prod_code",
        #             "prod_name",
        #             "risk_rating",
        #             "currency",
        #             "latest_nav_price",
        #             "3year_risk_return_ratio",
        #             "sharpe_ratio",
        #             "valuation_date",
        #             "cp_1m",
        #             "cp_3m",
        #             "cp_6m",
        #             "cp_ytd",
        #             "cp_1y",
        #             "cp_3y",
        #             "cp_5y",
        #             "cp_since_launch",
        #             "launch_date",
        #             "cyp_1",
        #             "cyp_2",
        #             "cyp_3",
        #             "cyp_4",
        #             "cyp_5"]
        # ws.append(ws_title)
        # for fund in fund_data:
        #     ws.append(fund)
        # wb.save(output_file)
        #
        # # delete all duplicated records:
        # cursor = cnx.cursor()
        # cursor.execute("""DELETE FROM t_fund_product WHERE data_source='FSM' and date(update_time)=curdate()""")
        # logging.info(unicode(cursor.rowcount) + ' FSM funds deleted')
        #
        # add_product = ("""INSERT INTO t_fund_product(prod_code, prod_name, risk_rating, currency, latest_nav_price,
        #                   3year_risk_return_ratio, sharpe_ratio, valuation_date, cp_1m, cp_3m, cp_6m, cp_ytd,
        #                   cp_1y, cp_3y, cp_5y, cp_since_launch, launch_date, cyp_1, cyp_2, cyp_3, cyp_4, cyp_5,
        #                   data_source, update_time)
        #                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        #                           %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'FSM', now())""")
        #
        # for fund in fund_data:
        #     cursor.execute(add_product, fund)
        #
        # logging.info(unicode(len(fund_data)) + ' FSM funds imported')
        # cursor.close()

    except Exception, err:
        logging.warning(unicode(sys.exc_info()[0]) + u':' + unicode(sys.exc_info()[1]))
        logging.exception("Exception:")
        logging.exception(traceback.format_exc())


def usage():
    print "-o [output_file] -v -h"


def version():
    print VERSION


if __name__ == '__main__':
    DB_NAME = 'zyq'

    try:
        # setup console parameters
        opts, args = getopt.getopt(sys.argv[1:], "hvo:", ["help", "version", "output="])

        output_destination = ''

        for op, value in opts:
            if op == '-o':
                output_destination = value
            elif op == '-h' or op == '--help':
                usage()
                exit(0)
            elif op == '-v' or op == '--version':
                version()
                exit(0)

        if not output_destination:
            output_file = "output/fund_" + LOCALTIME + ".xlsx"
        else:
            output_file = output_destination + "fund_FSM_" + LOCALTIME + ".xlsx"

        cnx = mysql.connector.connect(host='localhost', user='zyq', password='zyq', database=DB_NAME)
        cursor = cnx.cursor()

        logging.info('MYSQL connected.')

        # get_FSM_fund_classified_product()
        # get_jpm_fund_product()
        # ms_page = get_MS_fund_page_num()
        # get_MS_fund_product(ms_page)
        get_MS_fund_product_detail()
        set_MS_region()
        MS_export()

        # cursor.execute("""DELETE FROM t_fund_product WHERE data_source='MS' and date(update_time)=curdate()""")

        # total_page = get_MS_fund_page_num()
        # print "Total page: ", total_page

        # pool = Pool(3)
        # results = pool.map(get_MS_fund_product, range(total_page))

        # for page in range(total_page):
        #     get_MS_fund_product(page)

        cursor.close()

        cnx.commit()
        cnx.close()

    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)
            raise
